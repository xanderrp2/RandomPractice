import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter


tweets = load_workbook("/Users/USER/FILE/EXCEL_FILE.xlsx",data_only=True)
tweetsWS = tweets["Sheet1"]

checkingSamples = load_workbook("/Users/USER/FILE/EXCEL_FILE#2w.xlsx",data_only=True)
checkingSamplesWS = checkingSamples["BR"]

WB = Workbook()
WS = WB.active

# create 2D lists to hold cell data
rowtweets = []
rowSamples = []

# append cell data to lists
for row in range(1,58649):
    columnData = []
    for col in range(1,8):
            char = get_column_letter(col)
            columnData.append(tweetsWS[char+str(row)].value)
    rowtweets.append(columnData)

for row in range(1,148):
    CheckingData = []
    for col in range(1,4):
            char = get_column_letter(col)
            CheckingData.append(checkingSamplesWS[char+str(row)].value)
    rowSamples.append(CheckingData)

# merge data between two files
for i in rowtweets:
    for k in rowSamples:
        if str(i[0]) == str(k[2]):
            i[5] = k[0]


# print the new cell data into excel file
for b in range(0,len(rowtweets),1):
    WS.append(rowtweets[b])
WB.save("Vinies Tweets.xlsx")
